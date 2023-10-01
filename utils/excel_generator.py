import os
from datetime import datetime
from openpyxl import Workbook


class Report:
    """ Класс отчёта """

    def __init__(self, robot_serial_data: list, start_date: datetime, end_date: datetime):
        self.robot_serial_data = robot_serial_data
        self.start_date = start_date
        self.end_date = end_date

    # создание книги
    @staticmethod
    def _create_workbook(models: set) -> Workbook:
        wb = Workbook()
        sheets = [wb.create_sheet(model) for model in models]
        for sheet in sheets:
            sheet['A1'] = 'Модель'
            sheet['B1'] = 'Версия'
            sheet['C1'] = 'Количество за неделю'
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        return wb

    # заполнение листа
    @staticmethod
    def _fill_sheet(workbook: Workbook, sheet_name: str, order_info: dict) -> Workbook:
        model, version = order_info['robot_serial'].split('-')
        count = order_info['count']
        wb = workbook
        wb_sheet = wb[sheet_name]
        wb_sheet.append({'A': model, 'B': version, 'C': count})

        return wb

    # создание отчёта
    def _create_report(self) -> Workbook:
        models = set([model['robot_serial'].split('-')[0] for model in self.robot_serial_data])
        wb = self._create_workbook(models)
        for sheet in wb.sheetnames:
            for order_info in self.robot_serial_data:
                if order_info['robot_serial'].startswith(sheet):
                    wb = self._fill_sheet(workbook=wb, sheet_name=sheet, order_info=order_info)
        return wb

    # получение пути отчёта
    def get_path_with_report(self) -> str:
        start_date_str = self.start_date.strftime('%d.%m.%Y')
        end_date_str = self.end_date.strftime('%d.%m.%Y')
        file_name = f'report-{start_date_str}-{end_date_str}.xlsx'
        wb = self._create_report()
        report_path = os.path.join('media', file_name)
        wb.save(report_path)
        return report_path
