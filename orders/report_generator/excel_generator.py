from datetime import datetime
from openpyxl import Workbook
import os


# создание книги
def create_workbook(models: set) -> Workbook:
    wb = Workbook()
    sheets = [wb.create_sheet(model) for model in models]
    for sheet in sheets:
        sheet['A1'] = 'Модель'
        sheet['B1'] = 'Версия'
        sheet['C1'] = 'Количество за неделю'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    return wb


# заполнение листа
def fill_sheet(workbook: Workbook, sheet_name: str, order_info: dict) -> Workbook:
    model, version = order_info['robot_serial'].split('-')
    count = order_info['count']
    wb = workbook
    wb_sheet = wb[sheet_name]
    wb_sheet.append({'A': model, 'B': version, 'C': count})

    return wb


# создание отчёта
def create_report(robot_serial_data: list) -> Workbook:
    models = set([model['robot_serial'].split('-')[0] for model in robot_serial_data])
    wb = create_workbook(models)
    for sheet in wb.sheetnames:
        for order_info in robot_serial_data:
            if order_info['robot_serial'].startswith(sheet):
                wb = fill_sheet(workbook=wb, sheet_name=sheet, order_info=order_info)
    return wb


# получение ссылки на отчёт
def get_report(robot_serial_data: list, start_date: datetime, end_date: datetime) -> str:
    start_date_str = start_date.strftime('%d.%m.%Y')
    end_date_str = end_date.strftime('%d.%m.%Y')
    file_name = f'report-{start_date_str}-{end_date_str}.xlsx'
    wb = create_report(robot_serial_data)
    report_path = os.path.join('media', file_name)
    wb.save(report_path)
    return report_path
